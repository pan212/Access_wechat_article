from article_list import GetUser
from save_content import SaveContent
from get_detail import GetDetail
import openpyxl
import os
import re

import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)     # 忽略https证书提示

'''通过公众号内的文章获取到公众号的biz值，拼接出公众号主页链接'''
def get_article_link(url):
    '''输入公众号下任意一篇已发布的文章 短链接！！'''
    content = SaveContent().get_content(url)
    if content['content_flag'] == 1:
        print('正在获取微信公众号主页链接……')
        biz = re.search('__biz=(.*?)&', content['content']).group(1)
        names = re.search(r'var nickname.*"(.*?)".*', content['content']).group(1)  # 公众号名称
        main_url = ('https://mp.weixin.qq.com/mp/profile_ext?action=home&__biz=' + biz + '&scene=124#wechat_redirect')
        return {'link_flag': 1, 'main_url': main_url, 'names': names}
    else:
        print('未获取到文章内容，请检查链接是否正确')
        return {'link_flag': 0,}

'''
    得到公众号主页链接，粘贴到微信开始抓包，使用fiddler工具获取到请求的关键字值
    biz值：是微信公众号的标识码，每个公众号都有一个唯一的biz值
    uin值：就是user information，也就是特定微信用户的ID，通过fiddler抓到的包中获取
    key值：是一个动态参数
    pass_ticket值：微信登录之后返回的参数
'''
screen_text = '''请输入数字键！
    数字键1：获取公众号主页链接，输入公众号下任意一篇已发布的文章链接。
    数字键2：下载文章列表，每页文章有15篇
    数字键3：下载文章内容，自动下载文章列表中所有文章内容
    数字键4：下载‘文章列表.xlsx’中所有文章的 内容 + 细节信息（慎用！）
按其他任意键退出！\n'''

'''获取微信公众号列表，保存到指定文件夹'''
def save_article_list(pages, biz, uin, key, pass_ticket):
    '''pages为需要抓取的页数，一页有15篇文章'''
    getuser = GetUser()
    getuser.biz = biz
    getuser.uin = uin
    getuser.key = key
    getuser.pass_ticket = pass_ticket
    return getuser.seve_main(pages)


'''保存微信公众号列表内的文章正文内容'''
def save_article_content(save_path, save_img):
    '''
        save_path：文章保存路径，如：r'./data/国务院客户端/'
        读取文件：文章列表.xlsx
        创建文件：各微信文章内容.xlsx
        保存每篇文章到 save_path文件夹下
            文章文本内容保存到：各文章内容.xlsx
    '''
    # 输入文件地址，读取文章标题、文章链接
    column_title, column_url = [], []
    os.makedirs(save_path, exist_ok=True)   # 创建存储的文件夹
    if os.path.exists(save_path + '/文章列表.xlsx'):  # 检测文件是否存在，若不存在则新创建
        print("“文章列表.xlsx”文件存在")
        workbook = openpyxl.load_workbook(save_path + '/文章列表.xlsx')     # 打开：文章列表.xlsx
        sheet1 = workbook.active  # 获取活动工作表（通常是第一个工作表）
        column_title = [sheet1[f'C{i}'].value for i in range(2, sheet1.max_row + 1)]  # 读取C列数据，文章标题
        column_url = [sheet1[f'D{i}'].value for i in range(2, sheet1.max_row + 1)]  # 读取D列数据，文章链接
        print('获取到文章列表,一共' + str(len(column_title)) + '篇文章')
        print('开始保存各篇文章')
    else:
        print("“文章列表.xlsx”文件不存在")

    '''开始保存各篇文章'''
    GetUser().creat_content(save_path, '/微信文章内容.xlsx')  # 创建文章内容.xlsx，将文章内容进行保存操作
    sac = SaveContent()
    for i,j,z in zip(column_title, column_url, range(len(column_title))):
        # z为表头id，写入到'/微信文章内容.xlsx'时使用
        contents = sac.get_content(j)
        if contents['content_flag'] == 1:
            detail_time = sac.get_time(contents['content'])  # 获取文章发布时间信息
            texts = sac.get_texts(contents['content'])  # 列表形式的文章内容
            tittle = sac.get_title(contents['content'])
            print('正在下载文章：' + tittle)
            if save_img:
                sac.get_img(contents['content'], tittle, detail_time)  # 自动下载文章图片

            '''此处获取到文章信息，开始进行保存操作'''
            GetUser().write_content(save_path + '/微信文章内容.xlsx', z+1, detail_time[0],
                                    tittle, read_nums='', like_nums='', share_nums='',
                                    looking_nums='', links=j, comment='', comment_likes='', text_content=str(texts))
    if column_url:
        print('已完成保存操作，文章内容保存在：' + save_path, '微信文章内容.xlsx')
    else:
        print('文章列表为空，请先获取文章的链接信息，保存在‘文章列表.xlsx’内')


'''保存微信公众号文章的细节内容'''
def save_article_detail(save_path, biz, uin, key, pass_ticket, poc_sid):
    # 获取公众号名称
    # 建立获取细节信息类
    getdetail = GetDetail()
    getdetail.biz = biz
    getdetail.uin = uin
    getdetail.key = key
    getdetail.pass_ticket = pass_ticket
    getdetail.cookies['poc_sid'] = poc_sid
    passage_list = getdetail.get_message_new(1) # 获取一页公众号文章的链接
    # print(passage_list)
    names = ''
    if passage_list['m_flag'] == 1:
        if passage_list['passage_list'][0][0]:  # 存在至少一篇文章
            contents = getdetail.get_content(passage_list['passage_list'][0][2])
            poc_sid = contents['poc_sid']  # 更新poc_sid
            names = re.search(r'var nickname.*"(.*?)".*', contents['content']).group(1)  # 公众号名称
            save_path += names
            # 输入文件地址，读取文章标题、文章链接
            column_title, column_url = [], []
            os.makedirs(save_path, exist_ok=True)  # 创建存储的文件夹
            if os.path.exists(save_path + '/文章列表.xlsx'):  # 检测文件是否存在
                print("“文章列表.xlsx”文件存在")
                workbook = openpyxl.load_workbook(save_path + '/文章列表.xlsx')  # 打开：文章列表.xlsx
                sheet1 = workbook.active  # 获取活动工作表（通常是第一个工作表）
                column_title = [sheet1[f'C{i}'].value for i in range(2, sheet1.max_row + 1)]  # 读取C列数据，文章标题
                column_url = [sheet1[f'D{i}'].value for i in range(2, sheet1.max_row + 1)]  # 读取D列数据，文章链接
                print('\n获取到文章列表，将要下载' + str(len(column_title)) + '篇文章')
                print('开始获取文章细节信息，此处仅保留文字内容！！！！！')
            else:
                print("“文章列表.xlsx”文件不存在")

            '''开始保存所有文章'''
            getdetail.creat_content(save_path, '/微信文章内容.xlsx')  # 创建文章内容.xlsx，将文章内容进行保存操作
            for i, j, z in zip(column_title, column_url, range(len(column_title))):
                # print(getdetail.get_detail_new(i,j))
                (read_num, like_num, share_num, looking_num, comments,
                 like_numss, detail_time, title, texts) = getdetail.get_detail_new(j, i)

                '''此处获取到文章信息，开始进行保存操作'''
                print('正在保存：' + i)
                getdetail.write_content(save_path + '/微信文章内容.xlsx', z + 1, detail_time[0],
                                        title, read_num, like_num, share_num,
                                        looking_num, j, str(comments), str(like_numss), str(texts))
            if column_url:
                print('已完成保存操作，文章内容保存在：' + save_path, '/微信文章内容.xlsx')
                return {'name_flag': 1, 'names': names}
            else:
                print('文章列表为空，请先获取文章的链接信息，保存在‘文章列表.xlsx’内')
                return {'name_flag': 0}
        else:
            print('该公众号下没有文章')
            return {'name_flag': 0}
    else:
        print('文章信息获取失败，请检查后重试')
        return {'name_flag': 0}


if __name__=="__main__":
    # tst = r'File "D:\code\20240809_wechat_article\get_detail.py", line 144, in get_Alltypeif comment_id.group():^^^^^^^^^^^^^^^^AttributeError:'
    # print(re.search(r'eta_All(.*)d', tst))
    main_link = {   # 临时字典
        'names': '泰山风景名胜区', 'root_path': './data/',
        'biz': '', 'key': '', 'pass_ticket': '', 'uin': '', 'poc_sid': '',
        'main_url': '',  # 存放fiddler抓取的链接
    }
    print('欢迎使用，' + screen_text)
    while True:
        text = str(input('请输入功能数字：'))
        if text == '1':
            random_url = (input('（默认公众号主页链接为“国务院客户端，按回车键使用”）\n请输入公众号下任意一篇已发布的文章链接：') or
                          'https://mp.weixin.qq.com/s/WYPVC9AQzCWZIwaWFMl7Gw')
            main_link = get_article_link(random_url) or main_link
            print(main_link['names'] + '主页链接为：' + main_link['main_url'])
            print('将此链接👆👆👆粘贴发送到 ‘微信客户端-文件传输助手’')
            print('\n' + screen_text)
        elif text == '2':
            print('\n以下内容需要用到fiddler工具！！！！！\n'
                  '在微信客户端打开步骤1获取到的链接，\n'
                  '在fiddler中查看——主机地址为https://mp.weixin.qq.com，URL地址为：/mp/profile_ext?acti\n'
                  '选中此项后按快捷键：Ctrl+U，复制此网址到剪贴板\n'
                  '将该内容粘贴到此处👇👇👇')
            texts = input('请输入复制的链接：')
            main_link['biz'] = re.search('biz=(.*?)&', texts).group(1)
            main_link['uin'] = re.search('uin=(.*?)&', texts).group(1)
            main_link['key'] = re.search('key=(.*?)&', texts).group(1)
            main_link['pass_ticket'] = re.search('pass_ticket=(.*?)&', texts).group(1)
            pages = int(input('请输入需要下载的页数（一页有15篇文章，默认1页）：') or '1')
            names = save_article_list(pages, main_link['biz'], main_link['uin'], main_link['key'], main_link['pass_ticket'])
            if names['name_flag'] == 1:
                main_link['names'] = names['names']
            print('\n' + screen_text)
        elif text == '3':
            change_name, save_img = main_link['names'], ''
            print('默认保存路径：' + main_link['root_path'])
            text_names3 = input('检测到当前的微信公众号名称为：' + main_link['names'] +
                           '，是否更换公众号？\n是(输入任意值)，否(默认，直接按回车跳过)————(y/N)')
            if text_names3: # 输入新名称
                main_link['names'] = input('请输入微信公众号名称（例如：泰山风景名胜区）：')
                change_name = main_link['names']
            else:
                print('未更换公众号名称')
            save_path = main_link['root_path'] + change_name + '/'
            print('当前保存路径为：' + save_path)
            save_img = input('是否保存图片？是(输入任意值)，否(默认，直接按回车跳过)————(y/N)')
            save_article_content(save_path, save_img)
            print('\n' + screen_text)
        elif text == '4':
            print('\n以下内容需要用到fiddler工具！！！！！\n'
                  '在微信客户端打开步骤1获取到的链接，\n'
                  '在fiddler中查看——主机地址为https://mp.weixin.qq.com，URL地址为：/mp/profile_ext?acti\n'
                  '选中此项后按快捷键：Ctrl+U，复制此网址到剪贴板\n'
                  '将该内容粘贴到此处👇👇👇')
            texts = input('请输入复制的链接：')
            main_link['biz'] = re.search('biz=(.*?)&', texts).group(1)
            main_link['uin'] = re.search('uin=(.*?)&', texts).group(1)
            main_link['key'] = re.search('key=(.*?)&', texts).group(1)
            main_link['pass_ticket'] = re.search('pass_ticket=(.*?)&', texts).group(1)
            names = save_article_detail(main_link['root_path'], main_link['biz'],
                                        main_link['uin'], main_link['key'], main_link['pass_ticket'],
                                        main_link['poc_sid'])
            if names['name_flag'] == 1:
                main_link['names'] = names['names']
            print('\n' + screen_text)
        else:
            print('已成功退出！')
            break

